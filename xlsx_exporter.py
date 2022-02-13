import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from settings import JSON_FIELDS, OUT_XLSX


def dict_pre_writer(field):
    """
    Function prepares a dictionary for writing to the cell.
    :param field: dictionary from record of JSON object
    :return: string with sorted values
    """
    sorted_field = dict(sorted(field.items(), key=lambda x: -x[1]))
    res = ''
    for k, v in sorted_field.items():
        res += f'{v}% {k}, '
    return res[:-2]  # string without coma and space


def change_columns_width(ws, products):
    """
    Function change columns width according to the content
    :param ws: worksheet with needed columns
    :param products: JSON record - for rows counting
    :return: None.
    """
    for col_num in range(1, len(JSON_FIELDS) + 1):  # range with needed columns
        max_row_len = 3  # basic width
        for row_num in range(1, len(products) + 2):  # range with all rows
            cell = ws.cell(row=row_num, column=col_num)
            if len(str(cell.value)) > max_row_len:
                max_row_len = len(str(cell.value))
        ws.column_dimensions[get_column_letter(col_num)].width = max_row_len + 3


def write_xlsx(products, fname=OUT_XLSX):
    """
    Function saves the JSON object to a xlsx file
    :param products: JSON object
    :param fname: name of saving file
    :return: None
    """
    wb = Workbook()
    ws = wb.active
    products = json.loads(products)
    header_names = iter(JSON_FIELDS)
    # fill in  headers of table in worksheet
    for c in range(1, 9):
        ws.cell(row=1, column=c, value=next(header_names))

    # fill in records from JSON object to table rows
    for product in products:
        c = 1  # column number
        for field_name, field in product.items():
            if field_name != ws.cell(1, c).value:  # socks has no field type
                c += 1  # so going to the next column
            if isinstance(field, dict):
                field = dict_pre_writer(field)  # prepare data for writing in a cell
            ws.cell(row=product['id'] + 1, column=c, value=field)  # writing data in a cell
            c += 1

    change_columns_width(ws, products)
    wb.save(filename=fname)
